import openpyxl


def main():
    wb = openpyxl.load_workbook('D:\python\테스트.xlsx')
    print(type(wb))

    print(wb.get_sheet_names())

    sheet = wb.get_sheet_by_name('Sheet1')

    print(type(sheet))

    another_sheet = wb.active

    print(another_sheet)

    print('###########')

    # for i in range(1, 8, 2):
    #     print(i, sheet.cell(row=i, column=2).value)
    #
    # print('sheet.max_row : %s' % sheet.max_row)
    # print('sheet.max_column : %s' % sheet.max_column)
    #
    # print(sheet['A1'].value)
    # print(sheet.cell(row=3, column=3).value)

    for i in range(1, 10, 2):
        print(sheet.cell(row=i, column=3).value)


main()
